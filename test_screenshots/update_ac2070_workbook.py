from copy import copy

from openpyxl import load_workbook
from openpyxl.styles import Alignment


WORKBOOK = "AC2070_Medventory_HMU_Filled.xlsx"


def clear_range(ws, min_row, max_row=None, min_col=1, max_col=None):
    max_row = max_row or ws.max_row
    max_col = max_col or ws.max_column
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None


def copy_row_style(ws, src_row, dst_row, max_col=None):
    max_col = max_col or ws.max_column
    for col in range(1, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.number_format = src.number_format
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)


def write_rows(ws, start_row, rows, template_row=None):
    template_row = template_row or start_row
    clear_range(ws, start_row, ws.max_row, 1, ws.max_column)
    for row_idx, row in enumerate(rows, start=start_row):
        copy_row_style(ws, template_row, row_idx)
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="top",
                wrap_text=True,
            )


def apply_common_wrapping(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical="top",
                        wrap_text=True,
                    )


def set_widths(wb):
    widths = {
        "01_TongKetChucNang": {"A": 10, "B": 28, "C": 58, "D": 18, "E": 34, "F": 20, "G": 50},
        "02_TongKet_DiemMoi": {"A": 24, "B": 55, "C": 45, "D": 55, "E": 45, "F": 40},
        "03_PhanCong": {"A": 16, "B": 14, "C": 26, "D": 55, "E": 70, "F": 45, "G": 18, "H": 45, "I": 40},
        "04_KyThuat": {"A": 42, "B": 30, "C": 10, "D": 10, "E": 16, "F": 14, "G": 18, "H": 12},
        "05_Demo_KichBan": {"A": 30, "B": 38, "C": 10, "D": 55, "E": 55, "F": 34, "G": 50, "H": 30},
        "06_PhuLuc_CaNhan": {"A": 16, "B": 14, "C": 36, "D": 45, "E": 50, "F": 55, "G": 60, "H": 70, "I": 45, "J": 45, "K": 25},
        "99_GhiChu": {"A": 90, "B": 50},
    }
    for sheet, mapping in widths.items():
        ws = wb[sheet]
        for col, width in mapping.items():
            ws.column_dimensions[col].width = width


def main():
    wb = load_workbook(WORKBOOK)

    ws = wb["00_Checklist"]
    ws["A2"] = "Dựa theo hướng dẫn/ barem trong 'HD viết báo cáo và thang điểm.docx'. Cập nhật: 2026-05-05"
    for row in range(4, ws.max_row + 1):
        note = ws.cell(row, 8).value
        if isinstance(note, str):
            ws.cell(row, 8).value = note.replace(
                "responsive mobile, reset password, thông báo tự động",
                "responsive mobile, reset password, thông báo tự động, Admin/RBAC, reservation và lịch sử xuất kho",
            )

    ws = wb["01_TongKetChucNang"]
    rows_01 = [
        ("F01", "Đề nghị bổ sung hàng hóa", "Thủ kho lập phiếu dự trù vật tư theo bộ môn, có tìm kiếm vật tư, tự nạp số lượng hiện có/năm trước và gửi BGH xem xét.", "Done", "Long (BE) / Dũng (FE) / Trung (FE/Test)", "Hình 01-02", "POST /api/supp-forecast; GET /api/supp-forecast/previous; GET /api/inventory/materials"),
        ("F02", "Phê duyệt danh mục bổ sung", "BGH xem phiếu dự trù chờ duyệt, mở chi tiết trong modal, phê duyệt hoặc từ chối kèm ghi chú; hệ thống gửi thông báo kết quả.", "Done", "Long (BE) / Hùng (BE review) / Dũng (FE)", "Hình 03-05", "GET /api/supp-forecast/bgh/pending; GET /api/supp-forecast/bgh/processed; POST /api/supp-forecast/approve"),
        ("F03", "Nhập kho và bảo quản", "Thủ kho tạo phiếu nhập kho theo lô: vật tư, số lô, ngày sản xuất, hạn dùng, số lượng và đơn giá; có lịch sử và modal chi tiết.", "Done", "Long (BE) / Dũng (FE)", "Hình 06-08", "POST /api/receipts/create; GET /api/receipts/feed; GET /api/receipts/{id}/detail"),
        ("F04", "Phê duyệt cấp phát", "Cán bộ tạo phiếu xin lĩnh; lãnh đạo xem chi tiết, chỉnh quyền thao tác trong modal, duyệt/từ chối và ghi chú. Có cơ chế reservation tùy chọn.", "Done", "Hùng (Lead BE) / Dũng (FE) / Trung (Test)", "Hình 09-11", "POST /api/issue-requests/canbo/create; POST /api/issue-requests/{id}/approve; POST /api/issue-requests/{id}/reject"),
        ("F05", "Thực hiện xuất kho và bàn giao", "Thủ kho xem phiếu đủ điều kiện, preview lô xuất theo FEFO, xác nhận xuất kho và xem lịch sử phiếu xuất.", "Done", "Hùng (Lead BE) / Dũng (FE)", "Hình 12-14", "GET /api/issues/eligible-requests; POST /api/issues/create-from-issue-req; GET /api/issues/{id}/detail"),
        ("F06", "Đăng nhập / đăng ký / quên mật khẩu", "Trang auth có login/register/forgot/reset password; dashboard redirect về login khi chưa đăng nhập.", "Done", "Hùng (BE) / Dũng (FE)", "Hình 15-16", "POST /api/auth/login; POST /api/auth/register; POST /api/auth/forgot-password; POST /api/auth/reset-password"),
        ("F07", "Quản lý danh mục vật tư", "Thủ kho xem danh mục vật tư tồn kho, thêm vật tư mới, tìm kiếm và phân trang theo style chung.", "Done", "Long (BE) / Trung (DB/FE/Test) / Dũng (FE)", "Hình 17-18", "GET /api/inventory/materials; GET/POST /api/materials; GET /api/materials/categories"),
        ("F08", "Quản lý người dùng và phân quyền", "Admin quản lý tài khoản, phê duyệt/từ chối/xóa user trong modal chi tiết; phân quyền theo role và override theo user.", "Done", "Hùng (Lead BE) / Dũng (FE) / Trung (FE/Test)", "Hình 19-20", "GET /api/admin/users/all; POST /api/admin/users/{id}/approve; GET/POST /api/admin/rbac/*; GET/PUT /api/admin/settings/issue-req-auto-approve"),
        ("F09", "Thông báo, thống kê và kiểm tra tồn kho", "Header hiển thị thông báo nghiệp vụ; trang xuất kho dùng endpoint trả cả phiếu đủ/không đủ điều kiện; seed hiện tại chưa phát sinh phiếu không đủ điều kiện để chụp UI riêng.", "Done", "Long (BE) / Hùng (BE) / Trung (DB/Test) / Dũng (FE)", "Hình 21", "GET /api/notifications/my; POST /api/notifications/{id}/read; GET /api/issues/eligible-requests-with-reasons"),
    ]
    write_rows(ws, 4, rows_01, template_row=4)

    ws = wb["02_TongKet_DiemMoi"]
    rows_02 = [
        ("UI/UX", "Design system dùng chung trong frontend/src/components/dashboard-ui.css; bảng gọn hơn, có phân trang chung; modal chi tiết phiếu đồng bộ; responsive mobile card layout; toast/confirm rõ ràng.", "Chưa có dark mode; chưa tối ưu hoàn toàn cho màn hình rất nhỏ dưới 375px.", "Tái sử dụng CSS chung cho table, button, modal, pagination, badge trạng thái để giảm lệch giao diện giữa các trang.", "Một số component vẫn còn CSS riêng cho vùng nghiệp vụ đặc thù, cần tiếp tục rà soát nếu mở rộng.", "Tách thêm shared component Table/Pagination/DetailModal để giảm lặp JSX."),
        ("CSDL / Thiết kế DB", "PostgreSQL đầy đủ bảng vật tư, tồn kho theo lô, phiếu nhập/xuất/dự trù/xin lĩnh, users/roles/permissions/user_permissions, notifications, issue_reservations, system_settings. final_database.sql có seed reset test.", "Chưa dùng migration tool; password demo vẫn plaintext; index chưa tối ưu cho dữ liệu lớn.", "Có bảng system_settings để bật/tắt auto approve phiếu xin lĩnh; có issue_reservations để giữ tồn kho khi cần.", "ddl-auto/update và script reset phù hợp demo nhưng chưa phải quy trình production.", "Dùng Flyway/Liquibase, BCrypt, bổ sung index theo material_id/status/created_at."),
        ("API / Backend", "Spring Boot REST phân lớp Controller-Service-Repository; DTO pattern; validation; RBAC; 11 controllers; FEFO xuất kho; lịch sử nhập/xuất; reasons cho phiếu chưa đủ điều kiện xuất.", "Chưa có Swagger/OpenAPI; token/auth demo còn đơn giản qua localStorage + X-User-Id.", "RBAC theo role và override từng user; Admin tách riêng BGH, chỉ Admin quản lý user/quyền.", "X-User-Id có thể giả mạo nếu deploy thật; cần Spring Security/JWT.", "Thêm JWT, refresh token, audit log thao tác nhạy cảm."),
        ("Middleware / Auth / Validation", "CorsConfig cho phép FE 5173 gọi BE 8080; request gửi X-User-Id; controller kiểm tra quyền qua RbacService; App.jsx redirect về login khi chưa đăng nhập.", "Chưa có security filter chain chuẩn; chưa enforce HTTPS.", "Guard frontend tránh vào thẳng /dashboard khi đã logout; backend vẫn kiểm tra quyền từng API.", "Cần chuẩn hóa error response toàn backend để FE xử lý thống nhất hơn.", "Spring Security + JWT + global exception handler."),
        ("JS (FE) / Tương tác", "React 19 + Vite; React Router; hooks; controlled form; tìm kiếm real-time vật tư/bộ môn; phân trang chung; history modal; SweetAlert2 và react-hot-toast.", "Chưa có React Query/Zustand; URL API còn hardcode localhost trong nhiều component.", "Các luồng nghiệp vụ nhiều bước được xử lý trong một màn hình: tạo, lịch sử, xem chi tiết, duyệt/từ chối/xuất kho.", "Một số state/form logic còn nằm trực tiếp trong component lớn.", "Tách api client, hooks nghiệp vụ, dùng React Query và env config."),
        ("Tester / Demo", "Có script Playwright trong test_screenshots/capture_current_ui.py dùng skill webapp-testing để chụp lại ảnh UI hiện tại; ảnh pending/processed đã kiểm tra không trùng nội dung.", "Chưa có test tự động tích hợp CI cho toàn luồng nghiệp vụ.", "Ảnh demo được refresh theo code mới: Admin/RBAC, lịch sử xuất kho, phân trang, notification, auth redirect.", "Test hiện chủ yếu phục vụ demo/screenshot, chưa assert sâu business rule.", "Bổ sung test Playwright có assertion và test service backend."),
    ]
    write_rows(ws, 4, rows_02, template_row=4)

    ws = wb["03_PhanCong"]
    rows_03 = [
        ("Hùng", "[MSSV]", "Lead Backend", "Review luồng nghiệp vụ phê duyệt cấp phát, xuất kho, Admin/RBAC và notification để đảm bảo FE gọi đúng API.", "Thiết kế kiến trúc backend tổng thể Controller -> Service -> Repository; AuthController/AuthService; IssueReqController/IssueReqService; IssueController/IssueService; AdminController/RbacService; NotificationService; auto approve/reservation setting.", "CorsConfig; X-User-Id header auth demo; RBAC permission check tại controller/service; Spring @Valid; transaction cho nghiệp vụ xuất kho.", "Tốt", "github HungLe - permission and role, auto issue, special user, fix issue", "Giữ vai trò Lead Backend, không gán trách nhiệm chính frontend."),
        ("Trung", "[MSSV]", "Database / Frontend / Tester", "UI/UX shared style, dashboard-ui.css, EquipmentList, RBAC/admin UI hỗ trợ, responsive/pagination/table cleanup; kiểm thử và chụp ảnh demo.", "Thiết kế final_database.sql: materials, units, departments, inventory_card, receipt/issue/supp_forecast/issue_req header-detail, users/roles/permissions/user_permissions, notifications, issue_reservations, system_settings; hỗ trợ repository/query và seed data.", "Spring Data JPA schema support; React shared CSS; Playwright screenshot/testing workflow.", "Tốt", "github NLT22 - UI/UX overhaul, responsive, RBAC searchable role/user, test screenshots", "Đúng vai trò Database / Frontend / Tester."),
        ("Dũng", "[MSSV]", "Frontend Developer", "Xây dựng giao diện nghiệp vụ chính: ReplenishmentRequest, ForecastApproval, ReceiptPage, CreateIssueRequest, IssueRequestApproval, IssuePage, AuthForm/Forgot/Reset, MaterialSearchInput; kết nối API và xử lý modal/toast.", "Kết nối FE-BE bằng fetch, header X-User-Id, xử lý response JSON, validate client-side, cập nhật state sau action.", "React Router v7; SweetAlert2; react-hot-toast; fetch API; controlled form.", "Tốt", None, "Phụ trách chính frontend nghiệp vụ, không còn AddEquipment.jsx trong project hiện tại."),
        ("Long", "[MSSV]", "Backend Developer", "Hỗ trợ test luồng nhập kho, dự trù và danh mục vật tư từ phía giao diện.", "SuppForecastController/Service; ReceiptController/Service; MaterialController/Service; InventorySummaryController/Service; DepartmentController; UnitController; xử lý tồn kho theo lô và dữ liệu thống kê vật tư.", "Spring Data JPA; Spring @Valid; PostgreSQL JDBC; transaction nhập kho.", "Tốt", None, "Phụ trách backend dự trù, nhập kho, vật tư, thống kê tồn kho."),
    ]
    write_rows(ws, 4, rows_03, template_row=4)

    ws = wb["04_KyThuat"]
    rows_04 = [
        ("Đề nghị bổ sung hàng hóa", "Long / Dũng / Trung(Test)", "x", "x", "x", "x", "x", "x"),
        ("Phê duyệt danh mục bổ sung", "Long / Hùng(review) / Dũng", "x", "x", "x", "x", "x", "x"),
        ("Nhập kho và bảo quản", "Long / Dũng", "x", "x", "x", "x", "x", "x"),
        ("Phê duyệt cấp phát", "Hùng / Dũng / Trung(Test)", "x", "x", "x", "x", "x", "x"),
        ("Thực hiện xuất kho và bàn giao", "Hùng / Dũng", "x", "x", "x", "x", "x", "x"),
        ("Đăng nhập / quên mật khẩu / redirect", "Hùng / Dũng", "x", "x", "x", "x", "x", "x"),
        ("Quản lý danh mục vật tư", "Long / Trung / Dũng", "x", "x", "x", "x", "x", "x"),
        ("Admin quản lý người dùng và phân quyền", "Hùng / Dũng / Trung", "x", "x", "x", "x", "x", "x"),
        ("Thông báo, reservation và kiểm tra điều kiện xuất kho", "Hùng / Long / Trung / Dũng", "x", "x", "x", None, "x", "x"),
    ]
    write_rows(ws, 4, rows_04, template_row=4)

    ws = wb["05_Demo_KichBan"]
    rows_05 = [
        ("Đề nghị bổ sung hàng hóa", "Thủ kho lập phiếu dự trù vật tư cần bổ sung gửi BGH.", "1", "Thủ kho đăng nhập, vào tab \"Tạo phiếu dự trù\", chọn bộ môn bằng ô tìm kiếm và nhập danh sách vật tư.", "Form hiện vật tư đã chọn; số lượng hiện có/năm trước được nạp theo dữ liệu tồn kho và lịch sử.", "Hình 01 - supp_forecast_form.png", "GET /api/auth/departments/search; GET /api/inventory/materials; GET /api/supp-forecast/previous", None),
        ("Đề nghị bổ sung hàng hóa", "Thủ kho lập phiếu dự trù vật tư cần bổ sung gửi BGH.", "2", "Nhấn \"Gửi phiếu\" và chuyển sang lịch sử phiếu dự trù.", "Hệ thống lưu phiếu, hiển thị trạng thái và gửi thông báo tới BGH.", "Hình 02 - supp_forecast_sent.png", "POST /api/supp-forecast; GET /api/supp-forecast/my", None),
        ("Phê duyệt danh mục bổ sung", "BGH xem xét phiếu dự trù.", "1", "BGH vào tab \"Phê duyệt dự trù\" để xem danh sách chờ duyệt.", "Bảng pending hiển thị số liệu tổng quan, trạng thái và nút Xem.", "Hình 03 - forecast_pending_list.png", "GET /api/supp-forecast/bgh/pending; GET /api/supp-forecast/bgh/stats", None),
        ("Phê duyệt danh mục bổ sung", "BGH xem xét phiếu dự trù.", "2", "Nhấn \"Xem\" để mở chi tiết phiếu.", "Modal chi tiết dùng style chung, hiển thị thông tin phiếu, danh sách vật tư và nút duyệt/từ chối trong modal.", "Hình 04 - forecast_pending_detail.png", "GET /api/supp-forecast/{id}", None),
        ("Phê duyệt danh mục bổ sung", "BGH xem xét phiếu dự trù.", "3", "Chọn tab \"Đã xử lý\" sau khi duyệt hoặc từ chối.", "Danh sách processed khác pending, hiển thị phiếu đã duyệt/từ chối.", "Hình 05 - forecast_approved.png", "GET /api/supp-forecast/bgh/processed", None),
        ("Nhập kho và bảo quản", "Thủ kho ghi nhận hàng hóa nhập về kho.", "1", "Vào tab \"Nhập kho\", nhập nhà cung cấp, ngày nhập, lý do và danh sách vật tư theo lô.", "Form nhập kho hiển thị các trường số lô, ngày SX, hạn dùng, số lượng, đơn giá.", "Hình 06 - receipt_form.png", "GET /api/materials/search", None),
        ("Nhập kho và bảo quản", "Thủ kho ghi nhận hàng hóa nhập về kho.", "2", "Mở \"Lịch sử phiếu nhập\".", "Bảng lịch sử dùng phân trang chung, không dùng nút tải thêm.", "Hình 07 - receipt_history.png", "GET /api/receipts/feed", None),
        ("Nhập kho và bảo quản", "Thủ kho ghi nhận hàng hóa nhập về kho.", "3", "Nhấn \"Xem\" trên một phiếu nhập.", "Modal chi tiết phiếu nhập hiển thị tổng tiền và danh sách vật tư theo style chung.", "Hình 08 - receipt_saved.png", "GET /api/receipts/{id}/detail", None),
        ("Tạo phiếu xin lĩnh", "Cán bộ tạo phiếu xin lĩnh vật tư.", "1", "Cán bộ vào tab \"Tạo phiếu xin lĩnh\", thêm vật tư, mục đích sử dụng và gửi phiếu.", "Bảng nhập vật tư gọn, dùng search vật tư và phân loại vật tư mới/có sẵn.", "Hình 09 - create_issue_req.png", "POST /api/issue-requests/canbo/create", None),
        ("Phê duyệt phiếu xin lĩnh", "Lãnh đạo duyệt/từ chối phiếu xin lĩnh.", "1", "Lãnh đạo vào tab \"Phê duyệt phiếu xin lĩnh\".", "Danh sách chờ phê duyệt hiển thị badge compact, phân trang và nút Xem.", "Hình 10 - ira_preview.png", "GET /api/issue-requests/leader/pending", None),
        ("Phê duyệt phiếu xin lĩnh", "Lãnh đạo duyệt/từ chối phiếu xin lĩnh.", "2", "Chuyển sang lịch sử để xem phiếu đã xử lý.", "Danh sách lịch sử hiển thị phiếu đã duyệt/từ chối và lý do.", "Hình 11 - ira_approved.png", "GET /api/issue-requests/leader/processed", None),
        ("Xuất kho và bàn giao", "Thủ kho xuất kho theo phiếu đã duyệt.", "1", "Thủ kho vào tab \"Xuất kho\" xem phiếu đủ điều kiện xuất.", "Trang xuất kho hiển thị danh sách đủ điều kiện hoặc empty state, có nút tạo phiếu/lịch sử.", "Hình 12 - issue_preview.png", "GET /api/issues/eligible-requests", None),
        ("Xuất kho và bàn giao", "Thủ kho xuất kho theo phiếu đã duyệt.", "2", "Mở lịch sử phiếu xuất.", "Bảng lịch sử phiếu xuất dùng style chung và phân trang.", "Hình 13 - issue_done.png", "GET /api/issues/feed", None),
        ("Xuất kho và bàn giao", "Thủ kho xuất kho theo phiếu đã duyệt.", "3", "Nhấn \"Xem\" một phiếu xuất đã tạo.", "Modal chi tiết phiếu xuất hiển thị thông tin phiếu và danh sách vật tư xuất.", "Hình 14 - issue_history_detail.png", "GET /api/issues/{id}/detail", None),
        ("Đăng nhập / quên mật khẩu", "Người dùng truy cập hệ thống.", "1", "Mở trang đăng nhập.", "Trang login hiển thị form đăng nhập/đăng ký theo style hiện tại.", "Hình 15 - auth_login.png", "POST /api/auth/login; POST /api/auth/register", None),
        ("Đăng nhập / quên mật khẩu", "Người dùng truy cập hệ thống.", "2", "Chọn quên mật khẩu/reset password.", "Form reset password hiển thị trường email/mật khẩu và flow đổi mật khẩu.", "Hình 16 - auth_reset_password.png", "POST /api/auth/forgot-password; POST /api/auth/reset-password", None),
        ("Quản lý danh mục vật tư", "Thủ kho quản lý tồn kho.", "1", "Mở tab \"Danh sách vật tư\".", "Trang hiển thị stat cards, form thêm vật tư, bảng tồn kho gọn và phân trang.", "Hình 17 - equipment_list.png", "GET /api/inventory/materials; GET /api/materials/categories", None),
        ("Quản lý danh mục vật tư", "Thủ kho quản lý tồn kho.", "2", "Gõ từ khóa vào ô tìm kiếm vật tư.", "Bảng lọc theo mã/tên vật tư và giữ phân trang chung.", "Hình 18 - equipment_search.png", "GET /api/inventory/materials?keyword=...", None),
        ("Admin người dùng", "Admin quản lý tài khoản.", "1", "Admin vào tab \"Quản lý người dùng\".", "Danh sách user pending/approved, xem chi tiết trong modal, duyệt/từ chối/xóa/cập nhật quyền theo trạng thái.", "Hình 19 - admin_users.png", "GET /api/admin/users/all; POST /api/admin/users/{id}/approve", None),
        ("Admin phân quyền", "Admin cấu hình quyền hệ thống.", "1", "Admin vào tab \"Phân quyền vai trò\".", "RBAC chỉ thuộc Admin; BGH không còn quản lý người dùng/quyền.", "Hình 20 - rbac_permissions.png", "GET /api/admin/rbac/roles; POST /api/admin/rbac/roles/{id}/permissions", None),
        ("Thông báo", "Người dùng nhận thông báo nghiệp vụ.", "1", "Nhấn chuông thông báo ở header.", "Panel thông báo hiển thị danh sách thông báo, trạng thái chưa đọc và thời gian. Endpoint kiểm tra điều kiện xuất vẫn có trong backend; dữ liệu seed hiện tại chưa phát sinh UI phiếu không đủ điều kiện để chụp riêng.", "Hình 21 - notifications_dropdown.png", "GET /api/notifications/my; POST /api/notifications/{id}/read; GET /api/issues/eligible-requests-with-reasons", None),
    ]
    write_rows(ws, 4, rows_05, template_row=4)

    ws = wb["06_PhuLuc_CaNhan"]
    rows_06 = [
        ("Hùng", "[MSSV]", "Phê duyệt cấp phát, xuất kho, Admin/RBAC, reservation", "Spring Boot REST Controller; Spring Data JPA; X-User-Id header auth demo; RbacService; transaction service", "Spring Boot phù hợp REST API phân lớp; RBAC service tập trung giúp kiểm soát quyền; transaction đảm bảo xuất kho/trừ tồn/reservation nhất quán.", "Lead Backend: IssueReqController/Service, IssueController/Service, AdminController/RbacService, AuthController, NotificationService, auto approve/reservation setting.", "IssueReqController nhận request tạo/duyệt phiếu; service kiểm tồn/reservation theo setting; IssueController preview/xuất FEFO; AdminController quản lý user/quyền/setting.", "FE gửi request kèm X-User-Id -> Controller kiểm permission qua RbacService -> Service xử lý nghiệp vụ/JPA transaction -> Repository lưu DB -> NotificationService gửi thông báo -> trả DTO.", "HTTP 200 DTO nghiệp vụ hoặc HTTP 403 thiếu quyền / 400 dữ liệu không hợp lệ.", "Lead backend, review luồng nghiệp vụ và API contract cho FE.", None),
        ("Trung", "[MSSV]", "Database, shared UI và kiểm thử", "PostgreSQL schema; Spring Data JPA repository support; React shared CSS; Playwright webapp-testing", "PostgreSQL phù hợp dữ liệu tồn kho cần tính nhất quán; CSS chung giúp đồng bộ UI; Playwright giúp chụp/kiểm tra ảnh demo đúng trạng thái.", "Thiết kế final_database.sql; seed data reset test; dashboard-ui.css, phân trang/table/modal shared style; EquipmentList/RBAC UI hỗ trợ; tester.", "EquipmentList gọi GET /api/inventory/materials để lấy tồn kho tổng hợp; shared pagination xử lý phân trang client-side; script Playwright đăng nhập theo role và chụp từng màn hình.", "FE fetch /api/inventory/materials -> backend aggregate inventory_card/materials -> trả list/stat -> React lọc/tìm kiếm/phân trang -> render bảng; Playwright verify bằng ảnh.", "HTTP 200 danh sách vật tư/stat cards; ảnh test_screenshots/*.png dùng cho demo.", "Database / Frontend / Tester; đảm bảo final_database.sql và UI khớp code hiện tại.", "github: NLT22"),
        ("Dũng", "[MSSV]", "Frontend nghiệp vụ nhập kho, dự trù, xin lĩnh, phê duyệt, xuất kho", "React 19 hooks; React Router; fetch API; SweetAlert2; react-hot-toast; controlled form", "React SPA phù hợp nhiều role/tab; controlled form giữ dữ liệu nhất quán; SweetAlert2/toast cải thiện phản hồi người dùng.", "Xây dựng ReplenishmentRequest, ForecastApproval, ReceiptPage, CreateIssueRequest, IssueRequestApproval, IssuePage, AuthForm/Forgot/Reset, MaterialSearchInput.", "Component quản lý form state, gọi API, xử lý success/error, mở modal chi tiết và cập nhật danh sách sau thao tác.", "User submit form -> validate client -> fetch API đúng endpoint, ví dụ POST /api/receipts/create hoặc POST /api/issue-requests/canbo/create -> nhận JSON -> toast/modal -> refresh list.", "HTTP 200 DTO thành công hoặc HTTP 400/403 hiển thị SweetAlert/toast lỗi.", "Frontend Developer chính cho các luồng người dùng.", None),
        ("Long", "[MSSV]", "Backend nhập kho, dự trù bổ sung, vật tư và thống kê tồn kho", "Spring Boot REST; Spring Data JPA; Spring @Valid; PostgreSQL transaction", "Spring @Valid giảm validation thủ công; JPA transaction đảm bảo tạo receipt header/detail/inventory_card atomic; REST dễ tích hợp FE.", "ReceiptController/Service, SuppForecastController/Service, MaterialController/Service, InventorySummaryController/Service, DepartmentController, UnitController.", "ReceiptService tạo phiếu nhập và inventory_card theo lô; SuppForecastService tạo/duyệt phiếu dự trù; MaterialService quản lý vật tư/search; InventorySummaryService tổng hợp tồn kho.", "FE POST /api/receipts/create -> ReceiptController @Valid DTO -> ReceiptService @Transactional lưu receipt_header/detail + inventory_card -> trả DTO; FE GET /api/supp-forecast/previous để nạp dữ liệu năm trước.", "HTTP 200 Receipt/SuppForecast/Material DTO hoặc HTTP 400 validation error.", "Backend Developer cho nhập kho, dự trù, vật tư, thống kê.", None),
    ]
    write_rows(ws, 4, rows_06, template_row=4)

    ws = wb["99_GhiChu"]
    notes = [
        ("Điền dữ liệu vào các sheet 01-06 theo nội dung code hiện tại của Medventory-HMU.", None),
        ("Sheet 03 giữ đúng phân công: Hùng lead backend, Long backend, Dũng frontend, Trung database/frontend/tester.", None),
        ("Sheet 05 mapping ảnh demo với file trong thư mục test_screenshots/*.png; ảnh được chụp lại bằng skill .claude/skills/webapp-testing ngày 2026-05-05.", "test_screenshots/capture_current_ui.py"),
        ("Các ảnh dễ nhầm đã kiểm tra: forecast_pending_list khác forecast_approved; ira_preview khác ira_approved; notifications_dropdown đã refresh selector .dh-noti-btn.", None),
        ("Nguồn code chính: backend/src/main/java/com/backend, frontend/src/components, database/final_database.sql.", None),
    ]
    write_rows(ws, 4, notes, template_row=4)

    apply_common_wrapping(wb)
    set_widths(wb)
    wb.save(WORKBOOK)
    print(f"Updated {WORKBOOK}")


if __name__ == "__main__":
    main()
